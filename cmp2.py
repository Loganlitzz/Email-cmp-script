
from openpyxl import *
import xlsxwriter
filename3='newxl.xlsx'
filename2='orders.xlsx'

#wb=xlsxwriter.Workbook(filename2)

wb=load_workbook(filename2)
#ws=wb.active
print(wb.sheetnames)
'''
col_range=ws['A2:A14']
for a in ws.iter_rows(min_row=2, max_col=3, max_row=14,values_only=True):
    for value in a:
        print(value)
'''
wb.close()



